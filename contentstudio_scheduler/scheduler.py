#!/usr/bin/env python3
"""
ContentStudio Post QA Checker
Reads rows marked "Ready" from the Excel scheduler.
For each row, fetches the existing post from ContentStudio using the Post Link,
runs deep QA checks, updates Status to "Approved" or "QA Failed", and sends a
summary email.

Checks per platform:
  YouTube   — video filename, thumbnail filename, title (vs Word doc), copy (vs Word doc),
               title text visible on thumbnail image, schedule time, AI grammar
  Facebook  — video filename, copy (vs Word doc), schedule time, AI grammar
  LinkedIn  — video filename, copy (vs Word doc), schedule time, AI grammar
  Instagram — video filename, copy (vs Word doc), schedule time, AI grammar
  X         — thumbnail filename, title text on thumbnail, copy (vs Word doc),
               schedule time, AI grammar
"""

import base64
import os
import re
import smtplib
import urllib.parse
import requests
import pytz
import docx as docxlib
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv
import openpyxl
from anthropic import Anthropic

load_dotenv()

# ── Config ────────────────────────────────────────────────────────────────────
CS_API_KEY   = os.getenv("CONTENTSTUDIO_API_KEY")
CS_API_BASE  = "https://api.contentstudio.io/api/v1"
EXCEL_PATH   = os.getenv("EXCEL_PATH")
NOTIFY_EMAIL = os.getenv("NOTIFICATION_EMAIL", "jhammy@ringringmarketing.com")
SMTP_HOST    = os.getenv("SMTP_HOST", "smtp.office365.com")
SMTP_PORT    = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER    = os.getenv("SMTP_USER")
SMTP_PASS    = os.getenv("SMTP_PASS")

anthropic = Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))

# ── Workspace Mapping ─────────────────────────────────────────────────────────
WORKSPACE_MAP = {
    "Ring Ring Marketing":       "ring-ring-marketing",
    "RRM@home":                  "rrmathome",
    "Senior Care Marketing Max": "senior-care-marketing-max",
    "Home Care Post":            "home-care-post",
}

# ── Platform Rules ────────────────────────────────────────────────────────────
PLATFORM_RULES = {
    "youtube":   {"needs_video": True,  "needs_thumbnail": True,  "needs_title": True},
    "facebook":  {"needs_video": True,  "needs_thumbnail": False, "needs_title": False},
    "linkedin":  {"needs_video": True,  "needs_thumbnail": False, "needs_title": False},
    "instagram": {"needs_video": True,  "needs_thumbnail": False, "needs_title": False},
    "x":         {"needs_video": False, "needs_thumbnail": True,  "needs_title": False},
    "twitter":   {"needs_video": False, "needs_thumbnail": True,  "needs_title": False},
}

# ── Word Doc Section Mapping ──────────────────────────────────────────────────
BASE_SEARCH_DIRS = [
    r"D:\Ring Ring Marketing\Ring Ring Marketing\Messaging & Content Team - Personal Branding Videos",
    r"D:\Ring Ring Marketing\OneDrive - Ring Ring Marketing\RRM - In-House Marketing Initiatives - 2026",
]

PLATFORM_SECTION_MAP = {
    "facebook":  "Facebook and LinkedIn",
    "linkedin":  "Facebook and LinkedIn",
    "instagram": "Facebook and LinkedIn",
    "x":         "X",
    "twitter":   "X",
    "youtube":   "YouTube/ Website",
}

SECTION_HEADERS = ["Facebook and LinkedIn", "X", "YouTube/ Website"]
SKIP_LABELS     = {"title", "body", "tags", "yt link", "youtube tags"}

TIMEZONE_MAP = {
    "PST": "America/Los_Angeles", "PDT": "America/Los_Angeles",
    "MST": "America/Denver",      "MDT": "America/Denver",
    "CST": "America/Chicago",     "CDT": "America/Chicago",
    "EST": "America/New_York",    "EDT": "America/New_York",
}


# ── File Search ───────────────────────────────────────────────────────────────
def find_file(filename):
    name_lower = filename.strip().lower()
    for base in BASE_SEARCH_DIRS:
        if not os.path.isdir(base):
            continue
        for root, _, files in os.walk(base):
            for f in files:
                if f.lower() == name_lower:
                    return os.path.join(root, f)
                if os.path.splitext(f)[0].lower() == name_lower:
                    return os.path.join(root, f)
    return None


# ── Word Doc Parsing ──────────────────────────────────────────────────────────
def read_platform_copy(docx_path, platform):
    """Return (title, copy, hashtags) for the given platform from a Word doc."""
    target_section = PLATFORM_SECTION_MAP.get(platform.lower())
    if not target_section:
        raise ValueError(f"No section mapped for platform '{platform}'")

    doc = docxlib.Document(docx_path)
    paragraphs = [p.text.strip() for p in doc.paragraphs]

    in_section    = False
    next_is_title = False
    title_text    = ""
    copy_lines    = []
    hashtag_lines = []

    for para in paragraphs:
        stripped      = para.rstrip("─ \t–—")
        is_any_header = any(h.lower() in stripped.lower() for h in SECTION_HEADERS)
        is_target     = target_section.lower() in stripped.lower()

        if is_any_header:
            if is_target:
                in_section = True
            elif in_section:
                break
            continue

        if not in_section or not para:
            continue

        label = para.lower().rstrip(": ")
        if label in SKIP_LABELS:
            if label == "title":
                next_is_title = True
            continue

        if next_is_title:
            title_text    = para
            next_is_title = False
            continue

        if para.startswith("#"):
            hashtag_lines.append(para)
        else:
            copy_lines.append(para)

    copy     = "\n\n".join(l for l in copy_lines if l).strip()
    hashtags = " ".join(hashtag_lines).strip()
    return title_text, copy, hashtags


# ── ContentStudio API ─────────────────────────────────────────────────────────
def cs_headers():
    return {"X-API-Key": CS_API_KEY, "Content-Type": "application/json"}


def get_workspace_id(slug):
    resp = requests.get(f"{CS_API_BASE}/workspaces", headers=cs_headers(), timeout=15)
    resp.raise_for_status()
    for ws in resp.json().get("data", []):
        if ws.get("slug") == slug or ws.get("_id") == slug:
            return ws["_id"]
    raise ValueError(f"Workspace '{slug}' not found")


def extract_post_id(url):
    if not url:
        return None
    url = url.strip()
    if "plan_ids=" in url:
        return url.split("plan_ids=")[-1].split("&")[0]
    return url.rstrip("/").split("/")[-1]


def fetch_cs_post(workspace_id, post_id):
    """Scan paginated posts list to find the post with the given ID."""
    for page in range(1, 51):
        resp = requests.get(
            f"{CS_API_BASE}/workspaces/{workspace_id}/posts",
            headers=cs_headers(), params={"page": page}, timeout=15
        )
        resp.raise_for_status()
        data  = resp.json()
        posts = data.get("data", [])
        for post in posts:
            if post.get("id") == post_id:
                return post
        if not posts or page >= data.get("last_page", 1):
            break
    raise ValueError(f"Post {post_id} not found in ContentStudio (checked 50 pages)")


# ── Filename Helpers ──────────────────────────────────────────────────────────
_UUID_RE = re.compile(
    r'_[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$',
    re.IGNORECASE
)

def _cs_base_name(url):
    """Extract the original base filename from a ContentStudio storage URL (no UUID, no ext)."""
    path = urllib.parse.urlparse(url).path.split("?")[0]
    filename = urllib.parse.unquote(path.split("/")[-1])
    base = os.path.splitext(filename)[0]
    base = _UUID_RE.sub("", base)
    return base.lower().strip()


def _filename_matches(url, expected):
    """Return True if the ContentStudio URL filename matches the expected filename."""
    cs   = _cs_base_name(url)
    exp  = os.path.splitext(expected)[0].lower().strip()
    # Accept if either is a prefix of the other (ContentStudio truncates long names)
    short = cs if len(cs) < len(exp) else exp
    long  = exp if len(cs) < len(exp) else cs
    return long.startswith(short[:max(8, len(short))])


# ── AI Comparison Helpers ─────────────────────────────────────────────────────
def _compare_copy(cs_text, doc_copy, platform):
    """Return a list of issues if the ContentStudio copy doesn't match the Word doc copy."""
    if not doc_copy.strip():
        return []
    prompt = (
        f"Compare these two versions of a {platform} social media post.\n\n"
        f"VERSION A (in ContentStudio):\n{cs_text[:2000]}\n\n"
        f"VERSION B (from Word doc — expected):\n{doc_copy[:2000]}\n\n"
        "Are they essentially the same content? "
        "Minor formatting differences, extra hashtags, or 'INSERT LINK' placeholders are acceptable.\n"
        "Reply MATCH if the core content is the same.\n"
        "Reply MISMATCH: followed by a brief one-line explanation if they differ meaningfully."
    )
    msg = anthropic.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=200,
        messages=[{"role": "user", "content": prompt}],
    )
    result = msg.content[0].text.strip()
    if result.upper().startswith("MISMATCH"):
        explanation = result[8:].lstrip(": ").strip()
        return [f"Copy mismatch — {explanation}"]
    return []


def _check_title_on_image(image_url, expected_title):
    """
    Download the thumbnail and use Claude vision to verify the title text on it
    matches the expected title. Returns an issue string or None.
    """
    try:
        img_resp = requests.get(image_url, timeout=30)
        img_resp.raise_for_status()
        img_b64    = base64.standard_b64encode(img_resp.content).decode("utf-8")
        media_type = img_resp.headers.get("Content-Type", "image/jpeg").split(";")[0]
        if media_type not in ("image/jpeg", "image/png", "image/gif", "image/webp"):
            media_type = "image/jpeg"

        prompt = (
            f"Look at this thumbnail image.\n"
            f"Expected title: \"{expected_title}\"\n\n"
            "What title or heading text is displayed on the image?\n"
            "Reply MATCH if the text on the image matches the expected title (minor differences OK).\n"
            "Reply MISMATCH: followed by what you actually see on the image if it does not match."
        )
        msg = anthropic.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=200,
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {
                    "type": "base64", "media_type": media_type, "data": img_b64
                }},
                {"type": "text", "text": prompt},
            ]}],
        )
        result = msg.content[0].text.strip()
        if result.upper().startswith("MISMATCH"):
            explanation = result[8:].lstrip(": ").strip()
            return f"Thumbnail title mismatch — {explanation}"
    except Exception as e:
        return f"Could not check title on thumbnail: {e}"
    return None


# ── QA ────────────────────────────────────────────────────────────────────────
def run_qa(row, cs_post):
    issues   = []
    platform = row["platform"].lower()
    rules    = PLATFORM_RULES.get(platform)

    if not rules:
        issues.append(f"Unknown platform '{row['platform']}'")
        return issues

    # Pull content from the post
    common  = cs_post.get("common", {})
    content = common.get("content", {})
    cs_text = content.get("text", "").strip()
    media   = content.get("media", {})
    video   = media.get("video")
    images  = media.get("images", [])

    # Title from platform-specific override
    cs_title = content.get("title", "").strip()
    if not cs_title:
        for override in cs_post.get("overrides", []):
            if override.get("platform", "").lower() == platform:
                cs_title = override.get("content", {}).get("title", "").strip()
                if cs_title:
                    break

    # ── Read Word doc for expected values ────────────────────────────────────
    docx_filename = row.get("post_copy", "").strip()
    doc_title, doc_copy, doc_hashtags = "", "", ""
    if docx_filename:
        docx_path = find_file(docx_filename)
        if not docx_path:
            issues.append(f"Word doc not found: '{docx_filename}'")
        else:
            try:
                doc_title, doc_copy, doc_hashtags = read_platform_copy(docx_path, platform)
            except Exception as e:
                issues.append(f"Could not read Word doc: {e}")

    # For thumbnail title checks (YouTube and X share the same thumbnail)
    # Always read the YouTube section for the expected title on the image
    doc_yt_title = doc_title  # already YouTube title if platform==youtube
    if platform in ("x", "twitter") and docx_filename:
        docx_path = find_file(docx_filename)
        if docx_path:
            try:
                doc_yt_title, _, _ = read_platform_copy(docx_path, "youtube")
            except Exception:
                pass

    # ── Copy check ───────────────────────────────────────────────────────────
    if not cs_text:
        issues.append("Post copy is empty in ContentStudio")
    elif doc_copy:
        issues.extend(_compare_copy(cs_text, doc_copy, platform))

    # ── Video check ──────────────────────────────────────────────────────────
    if rules["needs_video"]:
        if not video:
            issues.append("Video is not attached")
        else:
            expected_video = row.get("video_file", "").strip()
            if expected_video:
                video_url = video.get("url", "") if isinstance(video, dict) else str(video)
                if not _filename_matches(video_url, expected_video):
                    cs_name = _cs_base_name(video_url)
                    issues.append(
                        f"Video mismatch — expected '{expected_video}', "
                        f"ContentStudio has '{cs_name}'"
                    )
                else:
                    print(f"    video OK: '{expected_video}'")

    # ── Thumbnail check ───────────────────────────────────────────────────────
    if rules["needs_thumbnail"]:
        # YouTube: custom thumbnail is stored in video.thumbnail (the overlay)
        # X/Twitter: thumbnail is a regular image in images[]
        if platform in ("youtube",):
            video_thumb_url = (video.get("thumbnail", "") if isinstance(video, dict) else "") or ""
            # Auto-generated thumbnails from ContentStudio have random hash filenames
            # A custom uploaded thumbnail will contain the original filename
            expected_thumb = row.get("image_filename", "").strip()
            if not video_thumb_url:
                issues.append("Thumbnail is not attached")
            elif expected_thumb and not _filename_matches(video_thumb_url, expected_thumb):
                cs_name = _cs_base_name(video_thumb_url)
                issues.append(
                    f"Thumbnail mismatch — expected '{expected_thumb}', "
                    f"ContentStudio has '{cs_name}'"
                )
            else:
                print(f"    thumbnail OK: '{expected_thumb or video_thumb_url}'")
                # Check title text on the thumbnail image
                check_title = doc_yt_title or cs_title
                if video_thumb_url and check_title:
                    print(f"    Checking title on thumbnail image...")
                    title_issue = _check_title_on_image(video_thumb_url, check_title)
                    if title_issue:
                        issues.append(title_issue)
                    else:
                        print(f"    thumbnail title OK")
        else:
            # X and other platforms: thumbnail is in images[]
            if not images:
                issues.append("Thumbnail is not attached")
            else:
                thumb_url = images[0] if isinstance(images[0], str) else images[0].get("url", "")
                expected_thumb = row.get("image_filename", "").strip()
                if expected_thumb:
                    if not _filename_matches(thumb_url, expected_thumb):
                        cs_name = _cs_base_name(thumb_url)
                        issues.append(
                            f"Thumbnail mismatch — expected '{expected_thumb}', "
                            f"ContentStudio has '{cs_name}'"
                        )
                    else:
                        print(f"    thumbnail OK: '{expected_thumb}'")
                # Check title text on image
                check_title = doc_yt_title or cs_title
                if thumb_url and check_title:
                    print(f"    Checking title on thumbnail image...")
                    title_issue = _check_title_on_image(thumb_url, check_title)
                    if title_issue:
                        issues.append(title_issue)
                    else:
                        print(f"    thumbnail title OK")

    # ── Title check (YouTube) ─────────────────────────────────────────────────
    if rules["needs_title"]:
        if not cs_title:
            if doc_title:
                issues.append(f"Title is missing — expected: '{doc_title}'")
            else:
                issues.append("Title is missing")
        elif doc_title and doc_title.lower() != cs_title.lower():
            issues.append(
                f"Title mismatch — expected: '{doc_title}', "
                f"ContentStudio has: '{cs_title}'"
            )
        else:
            print(f"    title OK: '{cs_title}'")

    # ── Schedule time check ───────────────────────────────────────────────────
    expected_at = row.get("scheduled_at")
    if expected_at:
        sched        = cs_post.get("scheduling", {})
        execute_time = sched.get("execute_time", "")
        if not execute_time:
            issues.append("Post has no scheduled time set in ContentStudio")
        else:
            try:
                cs_dt        = datetime.fromisoformat(execute_time.replace("Z", "+00:00"))
                diff_minutes = abs((cs_dt - expected_at).total_seconds()) / 60
                if diff_minutes > 10:
                    issues.append(
                        f"Schedule time mismatch — Excel expects "
                        f"{expected_at.strftime('%Y-%m-%d %H:%M UTC')}, "
                        f"ContentStudio shows {cs_dt.strftime('%Y-%m-%d %H:%M UTC')}"
                    )
                else:
                    print(f"    schedule OK: {cs_dt.strftime('%Y-%m-%d %H:%M UTC')}")
            except Exception:
                pass

    # ── AI grammar check ─────────────────────────────────────────────────────
    if not issues and cs_text:
        grammar_issues = _ai_grammar(cs_text, row["platform"])
        if grammar_issues:
            issues.extend(grammar_issues)
        else:
            print(f"    grammar OK")

    return issues


def _ai_grammar(copy, platform):
    prompt = (
        f"Review this {platform} social media post for a professional marketing agency:\n\n"
        f"\"{copy}\"\n\n"
        "Check ONLY for: spelling errors, grammar errors, or unprofessional language.\n"
        "If everything looks good, reply with exactly: PASS\n"
        "If there are issues, reply with: FAIL\n"
        "Then list each issue on its own line. Be brief."
    )
    msg = anthropic.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=300,
        messages=[{"role": "user", "content": prompt}],
    )
    result = msg.content[0].text.strip()
    if result.upper().startswith("FAIL"):
        return [line.strip("- ").strip() for line in result.splitlines()[1:] if line.strip()]
    return []


# ── Email ─────────────────────────────────────────────────────────────────────
def send_summary_email(approved, failed):
    total = len(approved) + len(failed)

    if not failed:
        subject = f"All {total} post(s) approved"
        body    = f"All {total} post(s) passed QA.\n\n"
        for r in approved:
            body += f"  OK  {r['industry']} | {r['platform']} — {r['schedule_date']} {r['schedule_time']} {r['timezone']}\n"
    else:
        subject = f"{len(approved)} approved, {len(failed)} need attention"
        body = ""
        if approved:
            body += f"{len(approved)} post(s) approved:\n"
            for r in approved:
                body += f"  OK  {r['industry']} | {r['platform']} — {r['schedule_date']} {r['schedule_time']} {r['timezone']}\n"
            body += "\n"
        body += f"{len(failed)} post(s) failed QA:\n\n"
        for entry in failed:
            r = entry["row"]
            body += f"  FAIL  {r['industry']} | {r['platform']} — {r['schedule_date']} {r['schedule_time']} {r['timezone']}\n"
            for issue in entry["issues"]:
                body += f"        - {issue}\n"
            body += "\n"
        body += "Fix the issues above, then set Status back to 'Ready' to re-check."

    msg = MIMEMultipart()
    msg["From"]    = SMTP_USER
    msg["To"]      = NOTIFY_EMAIL
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_USER, SMTP_PASS)
        server.send_message(msg)


# ── Excel Helpers ─────────────────────────────────────────────────────────────
def read_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [cell.value.strip() if cell.value else cell.value for cell in ws[1]]

    rows = []
    for i, excel_row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in excel_row]
        if not any(values):
            continue
        data = dict(zip(headers, values))
        rows.append({
            "row_num":        i,
            "industry":       str(data.get("Industry") or "").strip(),
            "platform":       str(data.get("Platform") or "").strip(),
            "post_copy":      str(data.get("Post Copy") or "").strip(),
            "video_file":     str(data.get("Video File Name") or "").strip(),
            "image_filename": str(data.get("Image Filename") or "").strip(),
            "yt_link":        str(data.get("YT Link") or "").strip(),
            "schedule_date":  data.get("Schedule Date"),
            "schedule_time":  data.get("Schedule Time"),
            "timezone":       str(data.get("Timezone") or "PST").strip(),
            "post_link":      str(data.get("Post Link for Checking") or "").strip(),
            "status":         str(data.get("Status") or "").strip(),
        })
    return wb, ws, rows, headers


def update_status(ws, row_num, headers, status):
    col = headers.index("Status") + 1
    ws.cell(row=row_num, column=col).value = status


def parse_datetime(date_val, time_val, tz_str):
    tz = pytz.timezone(TIMEZONE_MAP.get(tz_str.upper(), "America/Los_Angeles"))

    if isinstance(date_val, datetime):
        date_part = date_val.date()
    else:
        for fmt in ("%d-%b-%y", "%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                date_part = datetime.strptime(str(date_val).strip(), fmt).date()
                break
            except ValueError:
                continue
        else:
            raise ValueError(f"Unrecognised date format: {date_val}")

    if isinstance(time_val, datetime):
        time_part = time_val.time()
    else:
        for fmt in ("%I:%M %p", "%H:%M", "%I:%M%p", "%H:%M:%S"):
            try:
                time_part = datetime.strptime(str(time_val).strip(), fmt).time()
                break
            except ValueError:
                continue
        else:
            raise ValueError(f"Unrecognised time format: {time_val}")

    return tz.localize(datetime.combine(date_part, time_part)).astimezone(pytz.UTC)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 50)
    print("ContentStudio Post QA Checker")
    print("=" * 50)

    wb, ws, rows, headers = read_excel()
    ready = [r for r in rows if r["status"] == "Ready"]
    print(f"Found {len(ready)} post(s) marked Ready\n")

    workspace_cache = {}
    approved = []
    failed   = []

    for row in ready:
        label = f"{row['industry']} | {row['platform']} | Row {row['row_num']}"
        print(f"Checking: {label}")

        try:
            row["scheduled_at"] = parse_datetime(
                row["schedule_date"], row["schedule_time"], row["timezone"]
            )
        except Exception as e:
            issues = [f"Date/time error: {e}"]
            print(f"  FAIL {issues[0]}")
            update_status(ws, row["row_num"], headers, "QA Failed")
            failed.append({"row": row, "issues": issues})
            continue

        if not row["post_link"]:
            issues = ["Post Link for Checking is missing"]
            print(f"  FAIL {issues[0]}")
            update_status(ws, row["row_num"], headers, "QA Failed")
            failed.append({"row": row, "issues": issues})
            continue

        post_id = extract_post_id(row["post_link"])
        slug    = WORKSPACE_MAP.get(row["industry"])
        if not slug:
            issues = [f"Unknown industry '{row['industry']}'"]
            print(f"  FAIL {issues[0]}")
            update_status(ws, row["row_num"], headers, "QA Failed")
            failed.append({"row": row, "issues": issues})
            continue

        try:
            if slug not in workspace_cache:
                workspace_cache[slug] = get_workspace_id(slug)
            workspace_id = workspace_cache[slug]

            print(f"  Fetching post {post_id}...")
            cs_post = fetch_cs_post(workspace_id, post_id)

            issues = run_qa(row, cs_post)

            if issues:
                print(f"  FAIL ({len(issues)} issue(s))")
                for iss in issues:
                    print(f"      - {iss}")
                update_status(ws, row["row_num"], headers, "QA Failed")
                failed.append({"row": row, "issues": issues})
            else:
                print(f"  OK Approved")
                update_status(ws, row["row_num"], headers, "Approved")
                approved.append(row)

        except Exception as e:
            print(f"  FAIL Error: {e}")
            update_status(ws, row["row_num"], headers, "QA Failed")
            failed.append({"row": row, "issues": [str(e)]})

    if approved or failed:
        wb.save(EXCEL_PATH)
        print(f"\nExcel saved: {EXCEL_PATH}")
        try:
            send_summary_email(approved, failed)
            print(f"Summary email sent to {NOTIFY_EMAIL}")
        except Exception as e:
            print(f"Email failed (fix SMTP AUTH): {e}")

    print("\nDone.")


if __name__ == "__main__":
    main()
